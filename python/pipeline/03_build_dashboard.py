"""
Stage 4 - Build a ready-to-open Excel dashboard from the model's exports.

This is the "decision-ready" deliverable (brief Part 4) in a form anyone can
open without Power BI or Tableau. It reads the four datasets that
02_model.py writes and assembles ONE .xlsx with native Excel charts:

    - KPI tiles              (model_metrics.csv)     "how much do we trust it"
    - Locality ranking bar   (locality_ranking.csv)  "which areas are expensive"
    - Price-driver bar       (coefficients.csv)      "what moves rent"
    - Predicted-vs-actual    (dashboard_data.csv)    the trust chart (+45 deg line)

No new model, no new data - pure presentation of the one linear model's output.

Run AFTER 02_model.py:
    python python/pipeline/03_build_dashboard.py
Output:
    dashboard/rent_radar_dashboard.xlsx
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.properties import PageSetupProperties

ROOT = Path(__file__).resolve().parents[2]
CLEAN = ROOT / "data/clean"
OUT = ROOT / "dashboard/rent_radar_dashboard.xlsx"

INK = "1F2937"        # slate text
ACCENT = "2563EB"     # blue
OK = "16A34A"         # green  - in line
HIGH = "DC2626"       # red    - listed above (overpricing)
LOW = "F59E0B"        # amber  - listed below (underpricing)
GRID = "E5E7EB"


def _title(ws, cell, text, size=13):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=size, color=INK)


def _sheet_from_df(wb, name, df):
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # header styling
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)
    return ws


def main():
    metrics = pd.read_csv(CLEAN / "model_metrics.csv").set_index("metric")["value"]
    ranking = pd.read_csv(CLEAN / "locality_ranking.csv")
    drivers = pd.read_csv(CLEAN / "coefficients.csv")
    dash = pd.read_csv(CLEAN / "dashboard_data.csv")

    # Ranking: real areas only (>=2 listings so a single flat can't top the chart).
    rank = (ranking[ranking["n_listings"] >= 2]
            .sort_values("median_rent_per_sqft", ascending=False)
            .head(15)[["locality", "median_rent_per_sqft", "tier"]]
            .reset_index(drop=True))

    # Drivers: the 10 biggest levers by absolute effect, most positive first.
    drv = drivers.reindex(drivers["coef_log"].abs().sort_values(ascending=False).index)
    drv = drv.head(10)[["feature", "effect_pct"]].sort_values("effect_pct")

    # Predicted-vs-actual, split by verdict so each gets its own colour + a y=x line.
    pva = dash[["actual_rent", "predicted_rent", "pricing_flag"]].copy()
    lo, hi = 0, int(max(pva["actual_rent"].max(), pva["predicted_rent"].max()) * 1.02)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- data sheets (chart sources) ----
    ws_rank = _sheet_from_df(wb, "Ranking", rank)
    ws_drv = _sheet_from_df(wb, "Drivers", drv)
    ws_pva = _sheet_from_df(wb, "PredVsActual", pva)
    _sheet_from_df(wb, "AllListings", dash)          # full drill-down table

    # diagonal reference points for the trust chart (two points define the y=x line)
    ws_pva["F1"], ws_pva["G1"] = "diag_x", "diag_y"
    ws_pva["F2"], ws_pva["G2"] = lo, lo
    ws_pva["F3"], ws_pva["G3"] = hi, hi

    # ---- dashboard sheet ----
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    for col in "ABCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 11
    # Print/export clean: one landscape page wide, so a PDF or screenshot of the
    # sheet is submission-ready without manual page setup.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = "A1:P49"

    _title(ws, "A1", "Rent Radar - Mumbai  |  What should this flat cost?", size=16)
    ws["A2"] = ("One linear model, 5-fold cross-validated. Every prediction is "
                "out-of-fold (the model never saw that flat in training).")
    ws["A2"].font = Font(italic=True, color="6B7280", size=10)

    # KPI tiles
    tiles = [
        ("Model accuracy (CV R2)", f"{metrics['cv_r2']:.2f}", f"+/- {metrics['cv_r2_std']:.2f} across folds"),
        ("Typical error (MAE)", f"Rs {int(metrics['mae_rupees']):,}", "per month, out-of-fold"),
        ("Listings priced", f"{int(metrics['n_listings']):,}", f"{int(metrics['n_localities'])} localities"),
        ("Thin localities", f"{int(metrics['n_solo'])}", "single-listing (advisory only)"),
    ]
    col = 1
    for label, big, sub in tiles:
        c = ws.cell(row=4, column=col); c.value = label
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        v = ws.cell(row=5, column=col); v.value = big
        v.font = Font(bold=True, size=20, color=INK)
        v.alignment = Alignment(horizontal="center")
        s = ws.cell(row=6, column=col); s.value = sub
        s.font = Font(italic=True, size=9, color="6B7280")
        s.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 2)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 2)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 2)
        col += 4

    # Chart 1: locality ranking (bar)
    bar1 = BarChart(); bar1.type = "bar"; bar1.title = "Most expensive localities (median rent/sqft, >=2 listings)"
    data = Reference(ws_rank, min_col=2, min_row=1, max_row=len(rank) + 1)
    cats = Reference(ws_rank, min_col=1, min_row=2, max_row=len(rank) + 1)
    bar1.add_data(data, titles_from_data=True); bar1.set_categories(cats)
    bar1.legend = None; bar1.height = 9; bar1.width = 16
    ws.add_chart(bar1, "A8")

    # Chart 2: price drivers (bar)
    bar2 = BarChart(); bar2.type = "bar"; bar2.title = "What moves rent (effect on rent, %)"
    d2 = Reference(ws_drv, min_col=2, min_row=1, max_row=len(drv) + 1)
    c2 = Reference(ws_drv, min_col=1, min_row=2, max_row=len(drv) + 1)
    bar2.add_data(d2, titles_from_data=True); bar2.set_categories(c2)
    bar2.legend = None; bar2.height = 9; bar2.width = 16
    ws.add_chart(bar2, "I8")

    # Chart 3: predicted vs actual (scatter) - one series per verdict + a y=x line
    sc = ScatterChart(); sc.title = "Predicted vs actual rent (points on the line = model agrees)"
    sc.x_axis.title = "Actual rent (Rs/month)"; sc.y_axis.title = "Predicted rent (Rs/month)"
    sc.height = 11; sc.width = 20
    sc.x_axis.scaling.min = lo; sc.x_axis.scaling.max = hi
    sc.y_axis.scaling.min = lo; sc.y_axis.scaling.max = hi
    sc.x_axis.delete = False; sc.y_axis.delete = False

    flags = [
        ("In line with model", OK),
        ("Listed above model (check overpricing)", HIGH),
        ("Listed below model (possible underpricing)", LOW),
    ]
    n = len(pva)
    # rows are grouped? No - build per-flag references by filtering into helper columns.
    # Simplest robust route: one series over ALL points coloured neutral, plus the line.
    xref = Reference(ws_pva, min_col=1, min_row=2, max_row=n + 1)
    yref = Reference(ws_pva, min_col=2, min_row=2, max_row=n + 1)
    pts = Series(yref, xref, title="Listings")
    mk = Marker(symbol="circle", size=5)
    mk.graphicalProperties = GraphicalProperties(solidFill=ACCENT)
    pts.marker = mk
    pts.graphicalProperties.line.noFill = True
    sc.series.append(pts)

    # y = x reference line
    lx = Reference(ws_pva, min_col=6, min_row=2, max_row=3)
    ly = Reference(ws_pva, min_col=7, min_row=2, max_row=3)
    line = Series(ly, lx, title="Perfect prediction (y = x)")
    line.marker = Marker(symbol="none")
    line.graphicalProperties.line.solidFill = INK
    line.graphicalProperties.line.width = 18000
    sc.series.append(line)
    ws.add_chart(sc, "A27")

    # caption
    ws["A47"] = ("Read: dots below the line are listed ABOVE the model (possible overpricing); "
                 "dots above are possible underpricing. Tight for 1-3 BHK; fans out for 5+ BHK "
                 "and single-listing localities, where the model is advisory only.")
    ws["A47"].font = Font(italic=True, size=9, color="6B7280")
    ws.merge_cells("A47:P48")
    ws["A47"].alignment = Alignment(wrap_text=True, vertical="top")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Sheets: Dashboard, Ranking ({len(rank)}), Drivers ({len(drv)}), "
          f"PredVsActual ({n}), AllListings ({len(dash)})")


if __name__ == "__main__":
    main()
